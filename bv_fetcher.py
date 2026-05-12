#!/usr/bin/env python3
"""
B站 BV号 动态数据批量采集器
=============================
参考 寒棠 Daily (hantang-daily) 的实现思路：
  - 批量查询 API (单视频 view 接口 + 并发)
  - 每批 ~100个 BVID，多线程并发
  - 随机 User-Agent 防反爬
  - 失败的请求自动重试（指数退避）
  - 进度条 + CSV/Excel/JSON 输出

用法:
  python bv_fetcher.py -i bv_list.txt -o result.csv
  python bv_fetcher.py -i bv_list.txt -o result.xlsx
  python bv_fetcher.py -i bv_list.xlsx -o result.json

输入文件: txt/csv/xlsx，每行一个 BV 号或 B站视频链接
"""

import argparse
import csv
import json
import os
import random
import re
import ssl
import sys
import time
import urllib.request
import urllib.error
import urllib.parse
import sqlite3

# macOS Python 3.12+ SSL 证书兼容
_SSL_CTX = ssl._create_unverified_context()
from concurrent.futures import ThreadPoolExecutor, as_completed
from datetime import datetime
from pathlib import Path

# ============================================================
# 配置
# ============================================================

NUM_WORKERS = 10          # 并发线程数
MAX_RETRIES = 3           # 失败重试次数
BATCH_SIZE = 100          # 每个线程一次处理多少个 BV
MIN_SLEEP = 0.3           # 请求间隔下限（秒）

UA_LIST = [
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 "
    "(KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36",
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:109.0) Gecko/20100101 Firefox/119.0",
    "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/605.1.15 "
    "(KHTML, like Gecko) Version/17.1 Safari/605.1.15",
    "Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 "
    "(KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36",
    "BiliFetcher/1.0 (your-email@example.com)",
]

# ============================================================
# 工具函数
# ============================================================

def get_ua() -> str:
    return random.choice(UA_LIST)


def fmt(seconds: float) -> str:
    """秒 → 可读时长"""
    if seconds < 60:
        return f"{seconds:.0f}s"
    m, s = divmod(int(seconds), 60)
    h, m = divmod(m, 60)
    if h > 0:
        return f"{h}h{m:02d}m"
    return f"{m}m{s:02d}s"


# ============================================================
# 单视频查询 API: /x/web-interface/view?bvid=BVxxx
# ============================================================

def fetch_single(bvid: str) -> dict | None:
    """
    调 B站 view 接口获取单个视频的完整数据（包含 stat 统计）。
    重试 MAX_RETRIES 次，均失败则返回 None。
    """
    url = f"https://api.bilibili.com/x/web-interface/view?bvid={bvid}"

    for attempt in range(1, MAX_RETRIES + 1):
        try:
            time.sleep(MIN_SLEEP + random.uniform(0, 0.3))

            req = urllib.request.Request(url)
            req.add_header("User-Agent", get_ua())
            req.add_header("Referer", "https://www.bilibili.com/")
            req.add_header("Accept", "application/json, text/plain, */*")

            resp = urllib.request.urlopen(req, timeout=15, context=_SSL_CTX)
            raw = resp.read().decode("utf-8")
            data = json.loads(raw)

            if data.get("code") != 0:
                # 视频不存在/已删除/不可见 → 跳过
                return None

            return data.get("data")

        except urllib.error.HTTPError as e:
            if e.code == 503:
                wait = 2 ** attempt
                time.sleep(wait)
                continue
            elif e.code == 412:
                wait = 5 * attempt
                time.sleep(wait)
                continue
            else:
                return None
        except (urllib.error.URLError, OSError, json.JSONDecodeError):
            wait = 2 ** attempt
            time.sleep(wait)
            continue
        except Exception:
            return None

    return None


def fetch_batch_view(bvid_list: list) -> list:
    """串行获取一批 BVID 的 view 数据（给单线程内部用）"""
    results = []
    for bv in bvid_list:
        data = fetch_single(bv)
        if data is not None:
            results.append(flatten_view(data))
        else:
            pass  # 视频不存在或出错，跳过
    return results


def flatten_view(data: dict) -> dict:
    """将 view 接口返回的 data 扁平化"""
    stat = data.get("stat", {})
    owner = data.get("owner", {})
    return {
        "bvid": data.get("bvid", ""),
        "aid": data.get("aid", 0),
        "title": data.get("title", ""),
        "up_name": owner.get("name", ""),
        "up_mid": owner.get("mid", 0),
        "pubdate": data.get("pubdate", 0),
        "pubdate_str": datetime.fromtimestamp(data.get("pubdate", 0)).strftime("%Y-%m-%d %H:%M:%S"),
        "duration": data.get("duration", 0),
        "cover": data.get("pic", ""),
        "tid": data.get("tid", 0),
        "view": stat.get("view", 0),
        "danmaku": stat.get("danmaku", 0),
        "reply": stat.get("reply", 0),
        "favorite": stat.get("favorite", 0),
        "coin": stat.get("coin", 0),
        "share": stat.get("share", 0),
        "like": stat.get("like", 0),
    }


# ============================================================
# 批量 API: medialist/gateway/base/resource/infos (用 AID)
# ============================================================

# 先用 view 接口获取一版数据（含 AID），再对失败的用批处理补充


# ============================================================
# 主采集流程
# ============================================================

def collect(bvid_list: list) -> list:
    """
    多线程分组采集。
    - 分 N 组，每组 BATCH_SIZE 个 BV
    - 每组的串行 fetch 交给一个线程
    - 10 个线程并发
    """
    total = len(bvid_list)
    print(f"\n[{datetime.now().strftime('%H:%M:%S')}] 开始采集 {total} 个视频")

    # 分组
    batches = [bvid_list[i:i + BATCH_SIZE] for i in range(0, total, BATCH_SIZE)]
    print(f"  分批: {len(batches)} 组 × 每组 ≤{BATCH_SIZE} 个")
    print(f"  线程: {NUM_WORKERS}")
    print()

    all_records = []
    failed_bvids = []

    with ThreadPoolExecutor(max_workers=NUM_WORKERS) as pool:
        fut_map = {pool.submit(fetch_batch_view, batch): batch for batch in batches}

        done = 0
        total_batches = len(batches)
        for fut in as_completed(fut_map):
            batch = fut_map[fut]
            done += 1
            try:
                recs = fut.result()
                all_records.extend(recs)
                # 收集失败的 BV
                succeeded = {r["bvid"] for r in recs}
                for bv in batch:
                    if bv not in succeeded:
                        failed_bvids.append(bv)
            except Exception as e:
                failed_bvids.extend(batch)
                print(f"\n  ✗ 批次异常: {e}")

            # 进度条
            pct = done / total_batches * 100
            bar_sz = 30
            fill = int(bar_sz * done / total_batches)
            bar = "█" * fill + "░" * (bar_sz - fill)
            elapsed = time.time() - global_start
            eta = (elapsed / done) * (total_batches - done) if done > 0 else 0
            sys.stdout.write(
                f"\r  [{bar}] {done}/{total_batches} "
                f"({pct:.1f}%) · 已获取 {len(all_records)} 条 "
                f"· 耗时 {fmt(elapsed)} · 剩余 ≈{fmt(eta)}"
            )
            sys.stdout.flush()

    print()

    # 去重（保留最后一条）
    seen = set()
    deduped = []
    for r in reversed(all_records):
        if r["bvid"] not in seen:
            seen.add(r["bvid"])
            deduped.append(r)
    deduped.reverse()

    if failed_bvids:
        print(f"\n  ⚠ {len(failed_bvids)} 个视频获取失败")

    print(f"  ✓ 成功获取 {len(deduped)} 条")
    return deduped, list(set(failed_bvids))


# ============================================================
# 输入解析
# ============================================================

BV_RX = re.compile(r"BV[0-9A-Za-z]{10}")

def parse_bvids(path: str) -> list:
    """从文件提取 BV 号（去重），支持 txt/csv/xlsx 以及 SQLite 数据库"""
    seen = set()
    p = Path(path)
    ext = p.suffix.lower()

    # ---------- SQLite 数据库支持 ----------
    if ext in (".db", ".sqlite", ".sqlite3"):
        conn = sqlite3.connect(path)
        try:
            # 默认读取表 filtered_videos 中的 bvid 列（与 filter.py 输出一致）
            # 如果表不存在，尝试读取 videos 表（兼容 crawl_to_db 输出）
            table_name = None
            cursor = conn.execute(
                "SELECT name FROM sqlite_master WHERE type='table' AND name='filtered_videos'"
            )
            if cursor.fetchone():
                table_name = "filtered_videos"
            else:
                cursor = conn.execute(
                    "SELECT name FROM sqlite_master WHERE type='table' AND name='videos'"
                )
                if cursor.fetchone():
                    table_name = "videos"
            if not table_name:
                print(f"错误：数据库 {path} 中既没有 filtered_videos 表也没有 videos 表")
                return []
            cur = conn.execute(f"SELECT bvid FROM {table_name}")
            for row in cur:
                bv = row[0].strip()
                if bv:
                    seen.add(bv)
        finally:
            conn.close()
        result = list(seen)
        print(f"  输入数据库: {path}")
        print(f"  从表 {table_name} 读取到 {len(result)} 个唯一 BV 号")
        return result

    # ---------- 原有逻辑 ----------
    if ext in (".xlsx", ".xls"):
        try:
            import openpyxl
        except ImportError:
            print("需安装 openpyxl: pip install openpyxl")
            sys.exit(1)
        wb = openpyxl.load_workbook(p, read_only=True)
        ws = wb.active
        for row in ws.iter_rows(values_only=True):
            for cell in row:
                if cell is None:
                    continue
                for m in BV_RX.findall(str(cell)):
                    seen.add(m)
        wb.close()
    elif ext == ".csv":
        with open(p, "r", encoding="utf-8-sig") as f:
            for row in csv.reader(f):
                for cell in row:
                    for m in BV_RX.findall(cell):
                        seen.add(m)
    else:
        # 普通文本文件
        with open(p, "r", encoding="utf-8") as f:
            for line in f:
                line = line.strip()
                if not line:
                    continue
                for m in BV_RX.findall(line):
                    seen.add(m)

    result = list(seen)
    print(f"  输入文件: {path}")
    print(f"  解析到 {len(result)} 个唯一 BV 号")
    return result

# ============================================================
# 输出
# ============================================================

def save(records: list, path: str):
    if not records:
        print("无数据可保存")
        return
    ext = os.path.splitext(path)[1].lower()
    fields = list(records[0].keys())

    if ext == ".xlsx":
        try:
            import openpyxl
        except ImportError:
            print("openpyxl 未安装，改用 CSV")
            path = path.rsplit(".", 1)[0] + ".csv"
            ext = ".csv"

    if ext == ".xlsx":
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "视频数据"
        ws.append(fields)
        for r in records:
            ws.append([r.get(k, "") for k in fields])
        wb.save(path)
    elif ext == ".json":
        with open(path, "w", encoding="utf-8") as f:
            json.dump(records, f, ensure_ascii=False, indent=2)
    else:
        with open(path, "w", newline="", encoding="utf-8") as f:
            w = csv.DictWriter(f, fieldnames=fields)
            w.writeheader()
            w.writerows(records)

    print(f"\n  已保存 {len(records)} 条 → {path}")


# ============================================================
# CLI
# ============================================================

global_start = 0.0


def main():
    global global_start, NUM_WORKERS, BATCH_SIZE

    parser = argparse.ArgumentParser(
        description="B站 BV号 动态数据批量采集器",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
示例:
  python bv_fetcher.py -i bv.txt -o out.csv
  python bv_fetcher.py -i bv.xlsx -o out.xlsx
  python bv_fetcher.py -i bv.txt -o out.json

输入文件（txt/csv/xlsx）每行一个 BV 号或 B站链接。
        """,
    )
    parser.add_argument("-i", "--input", default="1.txt", help="输入文件")
    parser.add_argument("-o", "--output", default=None, help="输出文件 (.csv/.xlsx/.json)")
    parser.add_argument("-w", "--workers", type=int, default=10, help="线程数 (默认 10)")
    parser.add_argument("-s", "--size", type=int, default=100, help="每批数量 (默认 100)")
    parser.add_argument("--table", default=None, help="数据库中的表名（默认自动检测 filtered_videos / videos）")
    parser.add_argument("--column", default="bvid", help="表中存储 BV 号的列名（默认 bvid）")

    args = parser.parse_args()
    NUM_WORKERS = args.workers
    BATCH_SIZE = args.size

    if args.output is None:
        ts = datetime.now().strftime("%Y%m%d_%H%M")
        input_path = os.path.abspath(args.input)
        out_filename = f"{ts}.xlsx"
        args.output = os.path.join(os.path.dirname(input_path), out_filename)

    print()
    print(" ╔══════════════════════════════════════╗")
    print(" ║   B站 BV号 动态数据批量采集器        ║")
    print(" ║   参考 寒棠 Daily · 多线程并发       ║")
    print(" ╚══════════════════════════════════════╝")

    bvids = parse_bvids(args.input)
    if not bvids:
        print("未找到有效 BV 号")
        return

    print(f"  线程数: {NUM_WORKERS}, 每批: {BATCH_SIZE}")
    print()

    global_start = time.time()
    records, failed = collect(bvids)

    elapsed = time.time() - global_start
    print(f"\n[{datetime.now().strftime('%H:%M:%S')}] 采集完成")
    print(f"  总耗时: {fmt(elapsed)}")
    if records:
        print(f"  速率: {len(records) / elapsed:.1f} 条/秒")

    save(records, args.output)

    if failed:
        fail_path = args.output.rsplit(".", 1)[0] + "_failed.txt"
        with open(fail_path, "w") as f:
            for bv in failed:
                f.write(bv + "\n")
        print(f"  失败列表 → {fail_path} ({len(failed)} 个)")
    print()


if __name__ == "__main__":
    main()