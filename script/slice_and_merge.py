#!/usr/bin/env python3
"""
BV 号数据库切片、分发、合并工具
================================
用于将 filtered_videos.db 中的 BV 号平分成 3 份，
生成三个输入文件，供 bv_fetcher.py 在多台服务器上并行采集，
最后将三个结果文件合并，供 score_diff.py 使用。

用法：
  # 1) 切片：生成三个 txt 文件
  python slice_and_merge.py --mode slice --source filtered_videos.db --slice-prefix slice_

  # 2) 合并：将三台服务器生成的结果合并
  python slice_and_merge.py --mode merge --out merged_result.xlsx --result1 out1.csv --result2 out2.csv --result3 out3.csv

  # 3) 自动模式：本地切片 → 并行采集 → 合并（单机多进程加速）
  python slice_and_merge.py --mode auto --source filtered_videos.db --output merged_data.xlsx
"""

import argparse
import csv
import json
import os
import sqlite3
import sys
import subprocess
import time
from pathlib import Path
from typing import List, Dict, Any, Optional

# ========== 切片模式 ==========
def slice_db_to_files(source_db: str, slice_prefix: str = "slice_"):
    """
    从 SQLite 中读取所有 BV 号，均分成 3 份，每份写入一个 txt 文件。
    同时打印远程服务器上应运行的命令。
    """
    if not os.path.exists(source_db):
        print(f"错误：数据库 {source_db} 不存在", file=sys.stderr)
        sys.exit(1)

    conn = sqlite3.connect(source_db)
    cur = conn.execute("SELECT bvid FROM filtered_videos")
    rows = cur.fetchall()
    conn.close()

    bvids = [row[0] for row in rows]
    total = len(bvids)
    if total == 0:
        print("未找到任何 BV 号，退出")
        sys.exit(1)

    # 均分为 3 份
    size = total // 3
    parts = [
        bvids[0:size],
        bvids[size:2*size],
        bvids[2*size:]
    ]
    # 确保第三份包含所有剩余
    if len(parts[2]) < total - 2*size:
        parts[2] = bvids[2*size:]

    print(f"总 BV 数：{total}，分为 3 组：{len(parts[0])} / {len(parts[1])} / {len(parts[2])}")

    slice_files = []
    for i, bv_group in enumerate(parts):
        fname = f"{slice_prefix}{i}.txt"
        with open(fname, "w", encoding="utf-8") as f:
            for bv in bv_group:
                f.write(bv + "\n")
        slice_files.append(fname)
        print(f"已生成切片文件：{fname}")

    print("\n===== 请在每台服务器上分别执行以下命令 =====")
    for i, fname in enumerate(slice_files):
        cmd = f"python bv_fetcher.py -i {fname} -o result_slice_{i}.xlsx"
        print(f"服务器 {i+1}：{cmd}")
    print("==========================================\n")
    return slice_files


# ========== 合并模式 ==========
def load_data_from_file(file_path: str) -> Dict[str, Dict]:
    """加载 bv_fetcher 生成的任意格式文件，返回 {bvid: row_dict}"""
    ext = Path(file_path).suffix.lower()
    records = []

    if ext == ".csv":
        with open(file_path, "r", encoding="utf-8-sig") as f:
            reader = csv.DictReader(f)
            records = list(reader)
    elif ext in (".xlsx", ".xls"):
        try:
            import openpyxl
        except ImportError:
            sys.exit("需要安装 openpyxl：pip install openpyxl")
        wb = openpyxl.load_workbook(file_path, read_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        if rows:
            headers = [str(h) for h in rows[0]]
            for row in rows[1:]:
                rec = {headers[i]: row[i] for i in range(len(headers))}
                records.append(rec)
        wb.close()
    elif ext == ".json":
        with open(file_path, "r", encoding="utf-8") as f:
            records = json.load(f)
    else:
        raise ValueError(f"不支持的文件格式：{ext}")

    # 建立 bvid 索引
    db = {}
    for rec in records:
        bvid = rec.get("bvid")
        if bvid:
            db[bvid] = rec
    return db


def merge_results(result_files: List[str], output_path: str):
    """
    将多个 bv_fetcher 输出文件合并为一个文件（按 bvid 去重，保留最后出现的）
    """
    merged = {}
    for fpath in result_files:
        print(f"读取：{fpath}")
        data = load_data_from_file(fpath)
        merged.update(data)          # 后出现的会覆盖前面的，符合预期
        print(f"  添加 {len(data)} 条记录")

    records = list(merged.values())
    print(f"合并后共 {len(records)} 条唯一 BV 号")

    if not records:
        print("无有效数据，退出")
        return

    # 保存合并后的文件
    ext = Path(output_path).suffix.lower()
    if ext == ".csv":
        with open(output_path, "w", newline="", encoding="utf-8-sig") as f:
            writer = csv.DictWriter(f, fieldnames=records[0].keys())
            writer.writeheader()
            writer.writerows(records)
    elif ext in (".xlsx", ".xls"):
        try:
            import openpyxl
        except ImportError:
            sys.exit("需要安装 openpyxl：pip install openpyxl")
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(list(records[0].keys()))
        for rec in records:
            ws.append([rec.get(k, "") for k in records[0].keys()])
        wb.save(output_path)
    elif ext == ".json":
        with open(output_path, "w", encoding="utf-8") as f:
            json.dump(records, f, ensure_ascii=False, indent=2)
    else:
        print("未指定输出格式，默认保存为 CSV")
        with open(output_path + ".csv", "w", newline="", encoding="utf-8-sig") as f:
            writer = csv.DictWriter(f, fieldnames=records[0].keys())
            writer.writeheader()
            writer.writerows(records)

    print(f"合并结果已保存：{output_path}")


# ========== 自动模式（本地并行） ==========
def run_local_parallel(source_db: str, final_output: str, workers: int = 3, run_score_diff: bool = False, second_file: Optional[str] = None):
    """
    自动模式：
      1. 切片生成临时文件
      2. 并行运行 workers 个 bv_fetcher 进程
      3. 合并结果
      4. (可选) 调用 score_diff.py
    """
    # 1. 切片
    prefix = f"_temp_slice_"
    slice_files = slice_db_to_files(source_db, slice_prefix=prefix)
    output_parts = [f"{prefix}{i}_out.xlsx" for i in range(workers)]

    # 2. 并行运行 bv_fetcher
    print(f"\n开始并行采集，使用 {workers} 个进程...")
    processes = []
    for i, (input_file, output_file) in enumerate(zip(slice_files, output_parts)):
        cmd = [
            sys.executable, "bv_fetcher.py",
            "-i", input_file,
            "-o", output_file,
            "--quiet"          # 安静模式，避免输出混在一起
        ]
        print(f"启动进程 {i+1}：{' '.join(cmd)}")
        proc = subprocess.Popen(cmd)
        processes.append(proc)

    # 等待所有完成
    for proc in processes:
        proc.wait()

    print("所有采集进程已完成")

    # 3. 合并结果
    merge_results(output_parts, final_output)

    # 4. 清理临时文件（可选）
    for f in slice_files + output_parts:
        try:
            os.remove(f)
        except:
            pass

    # 5. 可选调用 score_diff.py
    if run_score_diff:
        if not second_file:
            print("错误：需要指定 --second-file 才能运行 score_diff.py")
            sys.exit(1)
        cmd = [
            sys.executable, "score_diff.py",
            second_file, final_output,
            "-o", f"{Path(final_output).stem}_diff.xlsx"
        ]
        print(f"\n调用 score_diff.py：{' '.join(cmd)}")
        subprocess.run(cmd)


# ========== 主入口 ==========
def main():
    parser = argparse.ArgumentParser(description="BV 号数据库切片、合并与并行采集辅助工具")
    parser.add_argument("--mode", choices=["slice", "merge", "auto"], required=True,
                        help="slice: 仅切片生成文件；merge: 合并三个结果；auto: 本地自动切片+并行采集+合并")
    parser.add_argument("--source", type=str, default="filtered_videos.db",
                        help="源 SQLite 数据库（默认 filtered_videos.db）")
    parser.add_argument("--slice-prefix", type=str, default="slice_",
                        help="切片文件前缀（默认 slice_）")
    parser.add_argument("--result1", type=str, help="服务器1的结果文件")
    parser.add_argument("--result2", type=str, help="服务器2的结果文件")
    parser.add_argument("--result3", type=str, help="服务器3的结果文件")
    parser.add_argument("--out", "--output", type=str, default="merged_result.xlsx",
                        help="合并后的输出文件路径")
    parser.add_argument("--second-file", type=str, help="用于 score_diff 的上一个时间点文件")
    parser.add_argument("--run-score-diff", action="store_true", help="自动模式下，合并后直接运行 score_diff.py")
    args = parser.parse_args()

    if args.mode == "slice":
        slice_db_to_files(args.source, args.slice_prefix)

    elif args.mode == "merge":
        if not args.result1 or not args.result2 or not args.result3:
            print("错误：merge 模式需要 --result1, --result2, --result3 三个参数")
            sys.exit(1)
        merge_results([args.result1, args.result2, args.result3], args.out)

    elif args.mode == "auto":
        run_local_parallel(
            source_db=args.source,
            final_output=args.out,
            workers=3,
            run_score_diff=args.run_score_diff,
            second_file=args.second_file
        )


if __name__ == "__main__":
    main()