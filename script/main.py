#!/usr/bin/env python3
# -*- coding: utf-8 -*-

"""
=================
整合采集、筛选、切片、云服务器创建、上传下载、合并、算分、删除实例等步骤。
支持自定义切片份数、服务器数量。

修复记录:
  1. 功能6(远程命令生成) — 修复 username 未定义导致的 NameError
  2. 全路径统一 — SCRIPT_DIR vs CWD 路径不混用
  3. 功能6 SSH命令改为利用 setup_and_run.sh（带虚拟环境 + 依赖安装）
  4. 功能5/7 硬编码脚本名改为 SCRIPTS 字典引用
  5. 功能8 os.listdir("..") 修复为当前目录
"""

import os
import sys
import subprocess
import sqlite3
import time
import json
import shutil
from pathlib import Path
from typing import List, Dict, Optional

# ---------- 全局配置 ----------
SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
WORKSPACE_DIR = os.path.join(SCRIPT_DIR, "..", "workspace")
DATA_DIR = os.path.join(SCRIPT_DIR, "..", "data")
os.makedirs(WORKSPACE_DIR, exist_ok=True)
# 脚本路径（确保与 main.py 同目录）
SCRIPTS = {
    "crawl_db": "crawl_to_db.py",
    "filter": "filter.py",
    "slice_merge": "slice_and_merge.py",
    "start_server": "startserver.py",
    "check_instance": "check_instance.py",
    "upload": "upload.py",
    "download": "download.py",
    "score_diff": "score_diff.py",
    "bv_fetcher": "bv_fetcher.py",
    "add_to_fav": "add_to_fav.py",
    "download_mp3": "download_mp3.py",
    "ncm_delete": "ncm_delete.py",
}

# 默认文件名（始终基于 SCRIPT_DIR 拼出完整路径，避免 CWD 不确定）
def _script_path(*parts: str) -> str:
    return os.path.join(SCRIPT_DIR, *parts)

INSTANCES_INFO = os.path.join(WORKSPACE_DIR, "instances_info.json")
DEFAULT_SOURCE_DB = os.path.join(WORKSPACE_DIR, "bilibili_videos.db")
DEFAULT_FILTERED_DB = os.path.join(WORKSPACE_DIR, "filtered_videos.db")
DEFAULT_SLICE_PREFIX = os.path.join(WORKSPACE_DIR, "slice_")

# SSH 默认凭据（与 startserver.py 中 LoginSettings.Password 一致）
DEFAULT_SSH_USER = "ubuntu"
DEFAULT_SSH_PASS = "X#7kPm$9qL@2wR&"
DEFAULT_REMOTE_DIR = "/home/ubuntu/"

# ---------- 辅助函数 ----------
def run_cmd(cmd: List[str], cwd=None) -> bool:
    """执行外部命令，返回是否成功"""
    print(f"\n>>> 执行命令: {' '.join(cmd)}")
    try:
        proc = subprocess.run(cmd, cwd=cwd or SCRIPT_DIR, check=False)
        if proc.returncode != 0:
            print(f"命令执行失败，返回码 {proc.returncode}")
            return False
        return True
    except Exception as e:
        print(f"执行异常: {e}")
        return False


def prompt_ssh_creds() -> tuple:
    """统一提示输入 SSH 凭据，返回 (username, password)"""
    user = input(f"SSH用户名 (默认{DEFAULT_SSH_USER}): ").strip() or DEFAULT_SSH_USER
    pwd = input(f"SSH密码 (默认{DEFAULT_SSH_PASS}): ").strip() or DEFAULT_SSH_PASS
    return user, pwd


def load_instances():
    """从 workspace 加载实例信息"""
    if not os.path.exists(INSTANCES_INFO):
        print(f"未找到实例信息文件: {INSTANCES_INFO}")
        print("请先执行菜单4「创建云服务器」")
        return []
    with open(INSTANCES_INFO, "r", encoding="utf-8") as f:
        data = json.load(f)
    instances = data.get("instances", [])
    if not instances:
        print("实例信息文件中没有实例数据")
        return []
    return instances


def get_bv_list_from_db(db_path: str) -> List[str]:
    """从 SQLite 数据库的 filtered_videos 或 videos 表中读取 BV 号列表"""
    if not os.path.exists(db_path):
        print(f"数据库不存在: {db_path}")
        return []
    conn = sqlite3.connect(db_path)
    # 尝试 filtered_videos 表，若无则回退到 videos
    tables = conn.execute(
        "SELECT name FROM sqlite_master WHERE type='table' AND name IN ('filtered_videos', 'videos')"
    ).fetchall()
    table_name = None
    for (tbl,) in tables:
        if tbl == "filtered_videos":
            table_name = tbl
            break
    if not table_name and tables:
        table_name = tables[0][0]
    if not table_name:
        conn.close()
        print("数据库中没有有效的视频表")
        return []
    cur = conn.execute(f"SELECT bvid FROM {table_name}")
    bvids = [row[0] for row in cur.fetchall()]
    conn.close()
    return bvids


def slice_bvids(bvids: List[str], num_slices: int, output_prefix: str) -> List[str]:
    """将 BV 号列表平分为 num_slices 份，生成 txt 文件到 workspace，返回文件路径列表"""
    total = len(bvids)
    if total == 0:
        print("没有 BV 号可切片")
        return []
    slice_size = total // num_slices
    remainder = total % num_slices
    files = []
    start = 0
    for i in range(num_slices):
        end = start + slice_size + (1 if i < remainder else 0)
        slice_bv = bvids[start:end]
        fname = f"{output_prefix}{i+1}.txt"
        full_path = os.path.join(WORKSPACE_DIR, fname)
        with open(full_path, "w", encoding="utf-8") as f:
            f.write("\n".join(slice_bv))
        files.append(fname)
        print(f"生成切片 {fname}: {len(slice_bv)} 个 BV")
        start = end
    return files


def merge_result_files(result_files: List[str], output_path: str):
    """合并多个结果文件，output_path 应指向 workspace"""
    if not result_files:
        print("没有结果文件可合并")
        return
    import openpyxl
    import csv

    merged = {}
    for fpath in result_files:
        ext = Path(fpath).suffix.lower()
        records = []
        if ext == ".csv":
            with open(fpath, "r", encoding="utf-8-sig") as f:
                reader = csv.DictReader(f)
                records = list(reader)
        elif ext in (".xlsx", ".xls"):
            wb = openpyxl.load_workbook(fpath, read_only=True)
            ws = wb.active
            headers = [cell.value for cell in ws[1]]
            for row in ws.iter_rows(min_row=2, values_only=True):
                rec = {headers[i]: val for i, val in enumerate(row) if i < len(headers)}
                records.append(rec)
            wb.close()
        else:
            print(f"跳过不支持的文件格式 {ext}: {fpath}")
            continue

        for rec in records:
            bvid = rec.get("bvid")
            if bvid:
                merged[bvid] = rec

    records_out = list(merged.values())
    if not records_out:
        print("无有效数据")
        return

    # 保存为 Excel
    wb = openpyxl.Workbook()
    ws = wb.active
    headers = list(records_out[0].keys())
    ws.append(headers)
    for rec in records_out:
        ws.append([rec.get(h, "") for h in headers])
    wb.save(output_path)
    print(f"合并完成，共 {len(records_out)} 条记录 -> {output_path}")


# ---------- 菜单功能实现 ----------
def menu_crawl_to_db():
    """1. 采集B站视频（crawl_to_db.py）"""
    print("\n▶ 开始采集B站视频（分区增量模式）")
    tid = input("请输入分区ID (默认30): ").strip()
    tid = tid if tid else "30"
    max_pages = input("最多爬取页数 (默认10000): ").strip()
    max_pages = max_pages if max_pages else "10000"
    since = input("手动指定水位线时间 (格式 YYYY-MM-DD HH:MM，直接回车则自动): ").strip()
    cmd = [sys.executable, SCRIPTS["crawl_db"], "--tid", tid, "--max-pages", max_pages]
    if since:
        cmd.extend(["--since", since])
    run_cmd(cmd)


def menu_filter():
    """2. 筛选视频（filter.py）"""
    print("\n▶ 开始筛选视频（时长、关键词、黑名单）")
    min_sec = input("最小时长(秒，默认120): ").strip()
    min_sec = min_sec if min_sec else "120"
    max_sec = input("最大时长(秒，默认420): ").strip()
    max_sec = max_sec if max_sec else "420"
    blacklist = input("黑名单文件 (默认blacklist.txt): ").strip()
    if blacklist:
        # 如果用户输入了文件名，转换为 workspace 下的绝对路径
        if not os.path.isabs(blacklist):
            blacklist = os.path.join(WORKSPACE_DIR, blacklist)
    else:
        blacklist = os.path.join(WORKSPACE_DIR, "blacklist.txt")

    cmd = [sys.executable, SCRIPTS["filter"], "--min", min_sec, "--max", max_sec, "--blacklist", blacklist]
    run_cmd(cmd)


def menu_slice():
    """3. 切片BV号（支持自定义份数）"""
    print("\n▶ 切片BV号")
    db_path = input(f"源数据库路径 (默认{DEFAULT_FILTERED_DB}): ").strip()
    if not db_path:
        db_path = DEFAULT_FILTERED_DB
    bvids = get_bv_list_from_db(db_path)
    if not bvids:
        print("未获取到BV号")
        return
    print(f"共 {len(bvids)} 个BV号")
    slices = input("请输入切割份数 (默认3): ").strip()
    slices = int(slices) if slices.isdigit() else 3
    prefix = input(f"输出文件前缀 (默认{DEFAULT_SLICE_PREFIX}): ").strip()
    if not prefix:
        prefix = DEFAULT_SLICE_PREFIX
    if not os.path.isabs(prefix):
        prefix = os.path.join(WORKSPACE_DIR, prefix)
    files = slice_bvids(bvids, slices, prefix)
    print(f"\n已生成 {len(files)} 个切片文件:")
    for f in files:
        print(f"  {f}")


def menu_create_servers():
    """4. 创建云服务器（支持自定义数量）"""
    print("\n▶ 创建腾讯云CVM实例")
    count = input("请输入要创建的实例数量 (默认3): ").strip()
    count = int(count) if count.isdigit() else 3
    cmd = [sys.executable, SCRIPTS["start_server"], "--count", str(count)]
    if not run_cmd(cmd):
        print("创建实例失败，请检查 startserver.py 是否支持 --count 参数。")
        print("提示：注意脚本内置了 Ubuntu 密码 X#7kPm$9qL@2wR&")
    else:
        print("实例创建请求已提交，请稍后检查 instances_info.json")


def menu_upload():
    """5. 上传数据到服务器（每个服务器上传对应的切片文件 + bv_fetcher.py + setup_and_run.sh）"""
    print("\n▶ 上传文件到云服务器")
    instances = load_instances()
    if not instances:
        return

    ips = [inst["PublicIp"] for inst in instances if inst.get("PublicIp")]
    if not ips:
        print("没有公网IP")
        return

    username, password = prompt_ssh_creds()
    remote_dir = input(f"远程目录 (默认{DEFAULT_REMOTE_DIR}): ").strip() or DEFAULT_REMOTE_DIR

    # 自动检测切片文件（按顺序 slice_1.txt, slice_2.txt, ...）
    slice_files = []
    for i in range(1, len(ips) + 1):
        candidate = os.path.join(WORKSPACE_DIR, f"slice_{i}.txt")
        if os.path.exists(candidate):
            slice_files.append(candidate)
        else:
            print(f"警告：未找到切片文件 slice_{i}.txt")
            print("请先执行菜单3「切片BV号」生成切片文件")
            return

    if len(slice_files) != len(ips):
        print(f"切片文件数量({len(slice_files)})与服务器数量({len(ips)})不匹配")
        return

    # 需要上传的脚本文件
    bv_fetcher_path = os.path.join(SCRIPT_DIR, "bv_fetcher.py")
    setup_script_path = os.path.join(SCRIPT_DIR, "setup_and_run.sh")
    for p, name in [(bv_fetcher_path, "bv_fetcher.py"), (setup_script_path, "setup_and_run.sh")]:
        if not os.path.exists(p):
            print(f"错误：{name} 不存在于脚本目录 {SCRIPT_DIR}")
            return

    print("\n开始上传...")
    for idx, (ip, slice_file) in enumerate(zip(ips, slice_files), start=1):
        print(f"\n>>> 上传到服务器 {idx} ({ip})")
        slice_basename = os.path.basename(slice_file)

        # 上传切片文件
        cmd_upload_slice = [
            sys.executable, _script_path(SCRIPTS["upload"]),
            "--local", slice_file,
            "--remote", os.path.join(remote_dir, slice_basename),
            "--host", ip,
            "--user", username,
            "--password", password,
        ]
        if not run_cmd(cmd_upload_slice):
            print(f"上传 {slice_basename} 失败，跳过该服务器")
            continue

        # 上传 bv_fetcher.py
        cmd_upload_bv = [
            sys.executable, _script_path(SCRIPTS["upload"]),
            "--local", bv_fetcher_path,
            "--remote", os.path.join(remote_dir, "bv_fetcher.py"),
            "--host", ip,
            "--user", username,
            "--password", password,
        ]
        if not run_cmd(cmd_upload_bv):
            print(f"上传 bv_fetcher.py 失败，可能影响远程执行")

        # 上传 setup_and_run.sh
        cmd_upload_setup = [
            sys.executable, _script_path(SCRIPTS["upload"]),
            "--local", setup_script_path,
            "--remote", os.path.join(remote_dir, "setup_and_run.sh"),
            "--host", ip,
            "--user", username,
            "--password", password,
        ]
        if not run_cmd(cmd_upload_setup):
            print(f"上传 setup_and_run.sh 失败，可能影响远程执行")

    print("\n✅ 所有上传任务完成")


def menu_remote_run():
    """6. 生成远程采集命令（SSH + 利用 setup_and_run.sh 一键执行）"""
    print("\n▶ 远程采集命令生成")
    instances = load_instances()
    if not instances:
        return

    ips = [inst["PublicIp"] for inst in instances if inst.get("PublicIp")]
    if not ips:
        print("无公网IP")
        return

    username, _ = prompt_ssh_creds()

    print("\n" + "=" * 60)
    print("请在各服务器上执行以下命令（一键安装依赖 + 采集）：")
    print("=" * 60)
    for idx, ip in enumerate(ips, 1):
        print(f"\n━━━ 服务器 {idx} (IP {ip}) ━━━")
        print(f"  ssh {username}@{ip}")
        print(f"  cd {DEFAULT_REMOTE_DIR}")
        print(f"  chmod +x setup_and_run.sh")
        print(f"  ./setup_and_run.sh slice_{idx}.txt")
    print("\n" + "=" * 60)
    print("💡 说明：")
    print("  - setup_and_run.sh 会自动安装 python3 + pip + openpyxl")
    print("  - 会在远程创建 venv 虚拟环境，避免依赖冲突")
    print(f"  - SSH密码（如未修改）: {DEFAULT_SSH_PASS}")
    print("  - 也可在 SSH 登录后直接复制命令执行")
    print("=" * 60)


def menu_download():
    """7. 从服务器下载结果文件（自动扫描远程目录下所有 .xlsx 和 _failed.txt）"""
    print("\n▶ 从服务器下载结果文件")
    instances = load_instances()
    if not instances:
        return

    ips = [inst["PublicIp"] for inst in instances if inst.get("PublicIp")]
    if not ips:
        print("没有公网IP")
        return

    username, password = prompt_ssh_creds()
    remote_dir = input(f"远程目录 (默认{DEFAULT_REMOTE_DIR}): ").strip() or DEFAULT_REMOTE_DIR
    local_dir = input(f"本地保存目录（默认{WORKSPACE_DIR}）: ").strip()
    if not local_dir:
        local_dir = WORKSPACE_DIR

    patterns = ["*.xlsx", "*_failed.txt"]
    print("\n开始下载...")

    for idx, ip in enumerate(ips, start=1):
        print(f"\n>>> 从服务器 {idx} ({ip}) 扫描文件")
        remote_files = []
        try:
            import paramiko
            ssh = paramiko.SSHClient()
            ssh.set_missing_host_key_policy(paramiko.AutoAddPolicy())
            ssh.connect(hostname=ip, username=username, password=password)

            # find with multiple -name patterns: no Python escapes needed
            find_parts = []
            for pat in patterns:
                find_parts.extend(['-name', pat, '-o'])
            find_parts.pop()  # remove trailing -o
            cmd_parts = ['find', remote_dir, '-maxdepth', '1', '-type', 'f', '\\('] + find_parts + ['\\)']
            find_cmd = ' '.join(cmd_parts)
            stdin, stdout, stderr = ssh.exec_command(find_cmd)
            remote_files = [line.strip() for line in stdout.readlines()]
            ssh.close()
        except Exception as e:
            print(f"  ❌ 连接或扫描失败: {e}")
            continue

        if not remote_files:
            print(f"  ⚠️ 未找到匹配的文件（模式：{patterns}）")
            continue

        print(f"  发现 {len(remote_files)} 个文件")
        for remote_path in remote_files:
            cmd_download = [
                sys.executable, _script_path(SCRIPTS["download"]),
                "--host", ip,
                "--user", username,
                "--password", password,
                "--remote", remote_path,
                "--local", local_dir,
            ]
            print(f"  下载: {os.path.basename(remote_path)}")
            run_cmd(cmd_download)

    print(f"\n✅ 下载完成，请检查目录: {local_dir}")


def menu_merge():
    """8. 合并结果文件（扫描 SCRIPT_DIR 下 .xlsx 文件）"""
    print("\n▶ 合并多个Excel文件")
    xlsx_files = [f for f in os.listdir(WORKSPACE_DIR) if f.endswith(".xlsx")]
    if not xlsx_files:
        print(f"未在 {WORKSPACE_DIR} 中找到任何 .xlsx 文件")
        return

    print(f"找到以下Excel文件（目录: {WORKSPACE_DIR}）:")
    for i, f in enumerate(xlsx_files, 1):
        size = os.path.getsize(os.path.join(WORKSPACE_DIR, f))
        print(f"  {i}. {f} ({size / 1024:.1f} KB)")

    selected = input("请输入要合并的文件序号（用空格分隔，默认全部）: ").strip()
    if selected:
        idxs = [int(x) - 1 for x in selected.split() if x.isdigit()]
        files_to_merge = [os.path.join(WORKSPACE_DIR, xlsx_files[i]) for i in idxs if i < len(xlsx_files)]
    else:
        files_to_merge = [os.path.join(WORKSPACE_DIR, f) for f in xlsx_files]

    if not files_to_merge:
        print("未选择任何文件")
        return

    out_name = input("输出文件名 (默认merged_result.xlsx): ").strip()
    out_name = out_name if out_name else "merged_result.xlsx"
    out_path = os.path.join(WORKSPACE_DIR, out_name)
    merge_result_files(files_to_merge, out_path)


def _pick_file_from_dir(dir_path: str, prompt: str, exclude: str = None) -> str:
    """列出目录下所有 .xlsx 文件，让用户选择，返回选中文件的完整路径。"""
    files = sorted(f for f in os.listdir(dir_path) if f.endswith(".xlsx"))
    if exclude:
        files = [f for f in files if os.path.join(dir_path, f) != exclude]
    if not files:
        print(f"{dir_path} 下没有 .xlsx 文件")
        return None
    print(f"\n可用文件:")
    for i, f in enumerate(files, 1):
        size = os.path.getsize(os.path.join(dir_path, f))
        print(f"  {i}. {f} ({size / 1024:.1f} KB)")
    sel = input(prompt).strip()
    if not sel.isdigit() or not (1 <= int(sel) <= len(files)):
        print("无效选择")
        return None
    return os.path.join(dir_path, files[int(sel) - 1])


def menu_score():
    """9. 计算得分（score_diff.py）"""
    print("\n▶ 计算增量得分")
    file1 = _pick_file_from_dir(DATA_DIR, "请选择旧文件 (输入序号): ")
    if not file1:
        return
    file2 = _pick_file_from_dir(DATA_DIR, "请选择新文件 (输入序号): ", exclude=file1)
    if not file2:
        return
    out = input("输出文件名 (默认自动生成): ").strip()
    cmd = [sys.executable, SCRIPTS["score_diff"], file1, file2]
    if out:
        SCORE_DIR = os.path.join(SCRIPT_DIR, "..", "score")
        os.makedirs(SCORE_DIR, exist_ok=True)
        out = os.path.join(SCORE_DIR, out)
        cmd.extend(["-o", out])
    run_cmd(cmd)


def menu_add_to_fav():
    """11. 将算分结果前 N 个视频加入收藏夹"""
    print("\n▶ 将算分结果加入B站收藏夹")
    score_files = [f for f in os.listdir(os.path.join(SCRIPT_DIR, "..", "score")) if f.endswith(".xlsx")]
    if not score_files:
        print(f"score/ 目录下没有算分文件，请先运行「9. 计算得分」")
        return
    print("可用算分文件:")
    for i, f in enumerate(score_files, 1):
        size = os.path.getsize(os.path.join(SCRIPT_DIR, "..", "score", f))
        print(f"  {i}. {f} ({size / 1024:.1f} KB)")
    sel = input("请选择 (输入序号): ").strip()
    if not sel.isdigit() or not (1 <= int(sel) <= len(score_files)):
        print("无效选择")
        return
    fpath = os.path.join(SCRIPT_DIR, "..", "score", score_files[int(sel) - 1])
    top_n = input("要收藏前几个视频? (默认 10): ").strip()
    top_n = top_n if top_n else "10"
    cmd = [sys.executable, SCRIPTS["add_to_fav"], fpath, "--top", top_n]
    run_cmd(cmd)


def menu_delete_instances():
    """10. 删除实例"""
    print("\n▶ 删除腾讯云实例")
    confirm = input("⚠️ 此操作将销毁所有实例，不可逆！是否继续？(yes/no): ").strip().lower()
    if confirm != "yes":
        print("取消删除")
        return
    run_cmd([sys.executable, SCRIPTS["check_instance"]])


def menu_download_mp3():
    """12. 下载MP3并上传网易云云盘"""
    print("\n▶ 下载算分结果前N个视频的MP3 → 上传网易云云盘 → 创建歌单")
    score_files = [f for f in os.listdir(os.path.join(SCRIPT_DIR, "..", "score")) if f.endswith(".xlsx")]
    if not score_files:
        print(f"score/ 目录下没有算分文件，请先运行「9. 计算得分」")
        return
    print("可用算分文件:")
    for i, f in enumerate(score_files, 1):
        size = os.path.getsize(os.path.join(SCRIPT_DIR, "..", "score", f))
        print(f"  {i}. {f} ({size / 1024:.1f} KB)")
    sel = input("请选择 (输入序号): ").strip()
    if not sel.isdigit() or not (1 <= int(sel) <= len(score_files)):
        print("无效选择")
        return
    fpath = os.path.join(SCRIPT_DIR, "..", "score", score_files[int(sel) - 1])
    top_n = input("要下载前几个视频? (默认 10): ").strip()
    top_n = top_n if top_n else "10"
    cmd = [sys.executable, SCRIPTS["download_mp3"], fpath, "--top", top_n]
    run_cmd(cmd)


def menu_ncm_delete():
    """13. 删除网易云歌单和云盘音乐"""
    print("\n▶ 删除网易云歌单和云盘音乐")
    confirm = input("⚠️ 将删除之前同步的歌单和云盘音乐，不可逆！是否继续？(yes/no): ").strip().lower()
    if confirm != "yes":
        print("取消删除")
        return
    run_cmd([sys.executable, SCRIPTS["ncm_delete"]])


def menu_exit():
    print("退出程序")
    sys.exit(0)


# ---------- 主菜单 ----------
def main():
    print(f"\n📁 脚本目录: {SCRIPT_DIR}")
    print(f"📁 实例信息: {INSTANCES_INFO}")
    print(f"📁 默认数据库: {DEFAULT_SOURCE_DB}")
    print(f"📂 工作空间: 所有生成的文件位于脚本目录下")

    while True:
        print("\n" + "=" * 50)
        print("         周刊工作流管理程序")
        print("=" * 50)
        print(" 1.  采集视频")
        print(" 2.  筛选视频")
        print(" 3.  切片BV号")
        print(" 4.  创建云服务器")
        print(" 5.  上传数据到服务器")
        print(" 6.  生成远程采集命令")
        print(" 7.  下载服务器结果")
        print(" 8.  合并结果文件")
        print(" 9.  计算得分")
        print("10.  删除所有服务器")
        print("11.  算分结果加入B站收藏夹")
        print("12.  下载MP3 → 上传网易云云盘 → 加入新歌单")
        print("13.  删除网易云歌单和云盘音乐")
        print(" 0.  退出")
        print("-" * 50)
        choice = input("请选择操作: ").strip()

        if choice == "1":
            menu_crawl_to_db()
        elif choice == "2":
            menu_filter()
        elif choice == "3":
            menu_slice()
        elif choice == "4":
            menu_create_servers()
        elif choice == "5":
            menu_upload()
        elif choice == "6":
            menu_remote_run()
        elif choice == "7":
            menu_download()
        elif choice == "8":
            menu_merge()
        elif choice == "9":
            menu_score()
        elif choice == "10":
            menu_delete_instances()
        elif choice == "11":
            menu_add_to_fav()
        elif choice == "12":
            menu_download_mp3()
        elif choice == "13":
            menu_ncm_delete()
        elif choice == "0":
            menu_exit()
        else:
            print("❌ 无效选项，请输入 0-13 之间的数字")
        input("\n按回车键继续...")

if __name__ == "__main__":
    main()
