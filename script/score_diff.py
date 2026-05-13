#!/usr/bin/env python3
"""
B站增量算分脚本
================
输入两个由 bv_fetcher.py 输出的数据文件（csv / xlsx / json），
以第二个文件（后一次时间节点）中的 BV 号为全集计算数据差值，并按照指定规则计算
最终得分 = 播放得分 + 互动得分 + 收藏得分 + 硬币得分 + 点赞得分

- 若某个 BV 号在两个文件中都存在 → 增量 = 时间点2 - 时间点1
- 若某个 BV 号仅存在于第二个文件 → 增量 = 时间点2（时间点1各项视为 0）

输出：
  - 包含增量、各项得分及最终得分的 Excel 文件
  - 文件名自动生成为 “时间点A-时间点B.xlsx”，也可通过 -o 指定

用法：
  python score_diff.py file_20260510_1200.csv file_20260511_1200.csv
  python score_diff.py data1.xlsx data2.xlsx -o diff_result.xlsx
"""

import argparse
import csv
import json
import os
import sys
import math
from pathlib import Path
from typing import Dict, List, Optional

# ---------- 文件读取 ----------
def load_data(path: str) -> Dict[str, dict]:
    """读取 bv_fetcher 输出的文件，返回 {bvid: 原始数据字典}"""
    p = Path(path)
    ext = p.suffix.lower()
    records = []

    if ext == ".csv":
        with open(p, "r", encoding="utf-8-sig") as f:
            reader = csv.DictReader(f)
            records = [row for row in reader]
    elif ext in (".xlsx", ".xls"):
        try:
            import openpyxl
        except ImportError:
            sys.exit("需安装 openpyxl: pip install openpyxl")
        wb = openpyxl.load_workbook(p, read_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            return {}
        headers = [str(h) for h in rows[0]]
        for row in rows[1:]:
            rec = {headers[i]: row[i] for i in range(len(headers))}
            records.append(rec)
        wb.close()
    elif ext == ".json":
        with open(p, "r", encoding="utf-8") as f:
            records = json.load(f)
    else:
        raise ValueError(f"不支持的文件格式: {ext}")

    # 以 bvid 为键建立索引
    db = {}
    for r in records:
        bvid = r.get("bvid")
        if bvid:
            db[bvid] = r
    return db


# ---------- 数值安全转换 ----------
def safe_float(val, default=0.0) -> float:
    """将字段转换为浮点数，转换失败则返回默认值"""
    try:
        return float(val)
    except (TypeError, ValueError):
        return default


# ---------- 算分核心 ----------
def calc_scores(delta: dict) -> dict:
    """
    输入增量字典，包含：
        view, danmaku, reply, favorite, coin, like
    返回各项得分及最终得分（保留两位小数）。
    """
    # 读取增量
    view = safe_float(delta.get("view", 0))
    danmaku = safe_float(delta.get("danmaku", 0))
    reply = safe_float(delta.get("reply", 0))
    fav = safe_float(delta.get("favorite", 0))
    coin = safe_float(delta.get("coin", 0))
    like = safe_float(delta.get("like", 0))

    # 防止除零
    if view == 0:
        return {
            "基础播放得分": 0.0,
            "修正A": 0.0,
            "修正B": 0.0,
            "修正C": 0.0,
            "修正D": 0.0,
            "播放得分": 0.0,
            "互动得分": 0.0,
            "收藏得分": 0.0,
            "硬币得分": 0.0,
            "点赞得分": 0.0,
            "最终得分": 0.0,
        }

    # --- 基础播放得分 ---
    if view > 10000:
        base_play = view * 0.5 + 5000
    else:
        base_play = view

    # --- 修正 D（最大值 1） ---
    if fav > coin:
        raw_d = (coin / view) * 25
    else:
        raw_d = (fav / view) * 25
    corr_d = min(raw_d, 1.0)
    corr_d = round(corr_d, 2)

    # --- 修正 A ---
    interaction = danmaku + reply
    numerator = base_play + fav
    denominator = base_play + fav + interaction * 20
    if denominator == 0:
        corr_a = 0.0
    else:
        corr_a = round((numerator / denominator) ** 2, 2)

    # --- 修正 B（最大值 50） ---
    if fav > coin * 2:
        raw_b = ((coin ** 2) / (view * fav)) * 1000
    else:
        raw_b = (fav / view) * 250
    corr_b = min(raw_b, 50.0)
    corr_b = round(corr_b, 2)

    # --- 修正 C（最大值 50） ---
    if coin > fav:
        if view != 0 and coin != 0:
            raw_c = ((fav ** 2) / (view * coin)) * 250
        else:
            raw_c = 0.0
    else:
        if view != 0:
            raw_c = (coin / view) * 250
        else:
            raw_c = 0.0
    corr_c = min(round(raw_c, 2), 50.0)

    # --- 各项得分 ---
    play_score = round(base_play * corr_d, 2)
    interact_score = round(interaction * corr_a * 15, 2)
    fav_score = round(fav * corr_b, 2)
    coin_score = round(coin * corr_c, 2)

    # 点赞得分
    if like > coin * 2:
        like_score = coin * 2
    else:
        like_score = like
    like_score = round(like_score, 2)

    final_score = round(play_score + interact_score + fav_score + coin_score + like_score, 2)

    return {
        "基础播放得分": round(base_play, 2),
        "修正A": corr_a,
        "修正B": corr_b,
        "修正C": corr_c,
        "修正D": corr_d,
        "播放得分": play_score,
        "互动得分": interact_score,
        "收藏得分": fav_score,
        "硬币得分": coin_score,
        "点赞得分": like_score,
        "最终得分": final_score,
    }


# ---------- 主函数 ----------
def main():
    parser = argparse.ArgumentParser(
        description="B站增量算分工具 - 计算两个时间点视频数据的增量得分"
    )
    parser.add_argument("file1",help="第一个时间点的数据文件（csv/xlsx/json）")
    parser.add_argument("file2",help="第二个时间点的数据文件（csv/xlsx/json）")
    parser.add_argument(
        "-o", "--output",
        default=None,
        help="输出文件路径（默认自动生成：时间点A-时间点B.xlsx）",
    )
    args = parser.parse_args()

    # 自动生成输出文件名
    if args.output is None:
        name1 = Path(args.file1).stem
        name2 = Path(args.file2).stem
        out_name = f"{name1}-{name2}.xlsx"
    else:
        out_name = args.output

    print(f"读取文件1: {args.file1}")
    data1 = load_data(args.file1)
    print(f"  获取 {len(data1)} 条记录")

    print(f"读取文件2: {args.file2}")
    data2 = load_data(args.file2)
    print(f"  获取 {len(data2)} 条记录")

    # 以第二个文件（后一次时间节点）中的所有 BV 号为全集
    all_bvids = set(data2.keys())
    only_in_2 = all_bvids - set(data1.keys())
    common = all_bvids & set(data1.keys())

    print(f"仅在时间点2新增的 BV 数量: {len(only_in_2)}")
    print(f"两个时间点共有的 BV 数量: {len(common)}")
    print(f"总计参与计算: {len(all_bvids)} 条")

    if not all_bvids:
        print("第二个文件中无任何 BV 号，无法计算。")
        return

    results = []
    for bvid in all_bvids:
        r2 = data2[bvid]

        # 若第一个文件中存在该 BV，则读取其数据，否则各项视为 0
        if bvid in data1:
            r1 = data1[bvid]
            v1_view = safe_float(r1.get("view", 0))
            v1_danmaku = safe_float(r1.get("danmaku", 0))
            v1_reply = safe_float(r1.get("reply", 0))
            v1_fav = safe_float(r1.get("favorite", 0))
            v1_coin = safe_float(r1.get("coin", 0))
            v1_like = safe_float(r1.get("like", 0))
            title = r2.get("title", "")
            up_name = r2.get("up_name", "")
            pubdate_str = r2.get("pubdate_str", "")
        else:
            v1_view = v1_danmaku = v1_reply = v1_fav = v1_coin = v1_like = 0.0
            title = r2.get("title", "")
            up_name = r2.get("up_name", "")

        # 时间点2数据
        v2_view = safe_float(r2.get("view", 0))
        v2_danmaku = safe_float(r2.get("danmaku", 0))
        v2_reply = safe_float(r2.get("reply", 0))
        v2_fav = safe_float(r2.get("favorite", 0))
        v2_coin = safe_float(r2.get("coin", 0))
        v2_like = safe_float(r2.get("like", 0))

        # 增量：对于新增视频，增量直接等于时间点2的值
        delta = {
            "view": v2_view - v1_view,
            "danmaku": v2_danmaku - v1_danmaku,
            "reply": v2_reply - v1_reply,
            "favorite": v2_fav - v1_fav,
            "coin": v2_coin - v1_coin,
            "like": v2_like - v1_like,
        }

        scores = calc_scores(delta)

        row = {
            "bvid": bvid,
            "title": title,
            "up_name": up_name,
            "pubdate_str": pubdate_str,
            # 时间点1 数据
            "view_1": v1_view,
            "danmaku_1": v1_danmaku,
            "reply_1": v1_reply,
            "favorite_1": v1_fav,
            "coin_1": v1_coin,
            "like_1": v1_like,
            # 时间点2 数据
            "view_2": v2_view,
            "danmaku_2": v2_danmaku,
            "reply_2": v2_reply,
            "favorite_2": v2_fav,
            "coin_2": v2_coin,
            "like_2": v2_like,
            # 增量
            "Δview": delta["view"],
            "Δdanmaku": delta["danmaku"],
            "Δreply": delta["reply"],
            "Δfavorite": delta["favorite"],
            "Δcoin": delta["coin"],
            "Δlike": delta["like"],
            # 算分过程
            **scores,
        }
        results.append(row)

    # 按最终得分降序排列
    results.sort(key=lambda x: x["最终得分"], reverse=True)

    # 保存为 Excel
    try:
        import openpyxl
    except ImportError:
        sys.exit("需安装 openpyxl 以输出 Excel: pip install openpyxl")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "增量算分"

    if results:
        headers = list(results[0].keys())
        ws.append(headers)
        for r in results:
            ws.append([r.get(h, "") for h in headers])
        wb.save(out_name)
        print(f"\n已保存结果至: {out_name}  ({len(results)} 条记录)")
    else:
        print("无有效数据可保存。")


if __name__ == "__main__":
    main()